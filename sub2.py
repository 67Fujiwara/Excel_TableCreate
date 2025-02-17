import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from pathlib import Path
from openpyxl import load_workbook
import string
from openpyxl.styles.alignment import Alignment
import openpyxl.styles
from tkinter import IntVar, Toplevel, ttk
import tkinter
import tkinter as tk
import pyperclip
import sys
import time
import tkinter.font as f
from openpyxl.worksheet.datavalidation import DataValidation
import sys

# print(sys.executable)


class WindowController:
    def __init__(self, master, title, bg_color, win_size):
        self.master = master
        self.master.title(title)
        self.master.config(bg=bg_color)
        self.master.geometry(win_size)
        self.ws_result = None
        self.ws = None
       
       
    def create_button(self, text, command, bg_color, font_color, xpos, ypos, width, height):
        button = tk.Button(self.master, text=text, command=command, bg=bg_color, fg=font_color)
        button.place(x=xpos, y=ypos, width=width, height=height)
        
        
    def create_labelframe(self, text, bg_color, font_color, xpos, ypos, width, height, label_title):
        self.labelframe = tk.LabelFrame(self.master, text=label_title, bg=bg_color, fg=font_color)
        self.labelframe.place(x=xpos-10, y=ypos-20, width=width+30, height=height+30)
        label = tk.Label(self.master, text=text, bg=bg_color, fg=font_color)
        label.place(x=xpos, y=ypos, width=width, height=height)
 
    
    def create_label(self, justify_pos, text, row, column, width):
        font1 = f.Font(family="Lucida Console", weight="bold", size=12, slant="italic")
        label = tk.Label(self.master, justify=justify_pos, text=text, bg="white")
        label["font"] = font1
        label.grid(row=row, column=column)
        entry = tk.Entry(self.master, width=width, justify=justify_pos)
        entry.grid(row=row , column=column + 1)
        return entry
    
    
    def create_label2(self, justify_pos, text, row, column):
        font = f.Font(family="Lucida Console", weight="bold", size=12, slant="italic")
        label = tk.Label(self.master, justify=justify_pos, text=text, bg="white")
        label["font"] = font
        label.grid(row=row, column=column)
    
    
    def create_label3(self, justify_pos, text, xpos, ypos, width, height):
        font = f.Font(family="Lucida Console", weight="bold", size=12, slant="italic")
        self.label = tk.Label(self.master, justify=justify_pos, text=text, bg="white", fg="green")
        self.label["font"] = font
        self.label.place(x=xpos, y=ypos, width=width, height=height)
        
    
    def create_entry(self, width, row, column):
        entry2 = tk.Entry(self.master, width=width)
        entry2.grid(row=row , column=column)
        return entry2
    
    def create_combobox(self, ws_count, char_count, justify_pos, ws_list, row, column):
        self.combobox = ttk.Combobox(self.master, 
                                height=ws_count, 
                                width=char_count, 
                                justify=justify_pos, 
                                state="readonly", 
                                values=ws_list,
                                cursor="arrow")
        self.combobox.grid(row=row, column=column)
        self.combobox.bind("<<ComboboxSelected>>", self.on_selection)
        
        
    def on_selection(self, event):
        self.ws_result = self.combobox.get()


    


class MainApp:
    def __init__(self, root):
        self.input_data() 
        self.root = root  
        self.sub1_isopen = False
        self.sub2_isopne = False
        self.bg_white = "white"
        self.font_black = "black"
        win_control = WindowController(root, "create_table", self.bg_white, "360x150")
        check_button = win_control.create_button("check", self.open_win1, self.bg_white, self.font_black, 40, 20, 100, 50)
        value_button = win_control.create_button("value", self.open_win2, self.bg_white, self.font_black, 210, 20, 100, 50)
        win_control.create_labelframe(file_path, self.bg_white, self.font_black, 20, 100, 300, 20, "File_Path")
   

    def input_data(self):
        self.wb = load_workbook(file_path, data_only=True)
        self.ws_list = self.wb.sheetnames
    
    
    def column_to_number(self, column):
        number = 0
        for c in column.upper():
            number = number * 26 + (ord(c) - ord('A') + 1)
        return number
    
        
    def input_check_table(self):
        # if self.sub1_isopen == True:
        #     self.ws = self.wb[self.sub1_control.ws_result]
        #     print('sub1')
        # elif self.sub2_isopne == True:
        #     print('sub2 ON')
        #     self.ws = self.wb[self.sub2_control.ws_result]
        #     print('sub2')
        # else:
        #     pass
        
        self.ws = self.wb[self.sub2_control.ws_result]
        self.wb.active = self.ws
        self.ref_column = self.getAlphabet_cell    #基準横軸
        self.ref_row = self.getNumber_cell     #基準縦軸
        self.deadDay_row = self.ref_row
        self.task_row = self.ref_row + 1
        
        self.reference_cell = [self.ref_column + str(self.ref_row)]
        self.input_columns = self.getNumber_column   #横軸   
        self.input_rows = self.getNumber_row       #縦軸
        self.cell_employeeName = 1   #社員名用ｾﾙ
        # self.cell_employeeNumber = 1 #社員ｺｰﾄﾞ用ｾﾙ
        self.cell_task = self.deadDay_row + 1  #期限用ｾﾙ + ﾀｽｸ用ｾﾙ  
        self.cell_achievement_rate = self.cell_employeeName +1  #達成率
        self.cell_target_value = self.cell_achievement_rate + 1  #目標金額

        
        self.column_number = self.column_to_number(self.ref_column)
        # if self.sub1_isopen:
        #     self.sub1_create_table()
        # elif self.sub2_isopne:
        #     self.sub2_create_table()
        # else:
        #     print('not sub_window control')
        #     pass
        
        self.sub2_create_table()
       
        return self.ref_column
     
     
    def open_win1(self):
        self.sub1_isopen = True
        self.check_sub = self.create_sub1()

    def open_win2(self):
        self.sub2_isopen = True
        self.value_sub = self.create_sub2()


    def create_sub1(self):
        sub_win1 = Toplevel(self.root)
        self.sub1_control = WindowController(sub_win1, "check_table", self.bg_white, "250x200")
        self.sub1_control.create_label2("center", "シート", 0, 0, )
        self.sub1_control.create_combobox(len(self.ws_list), 10, "center", self.ws_list, 0, 1)
        self.alphabet_cell = self.sub1_control.create_label("center", "基準セル", 1, 0, 5)
        self.number_cell = self.sub1_control.create_entry(5, 1, 2 )
        self.number_column = self.sub1_control.create_label("center", "列数", 2, 0, 10)
        self.number_row = self.sub1_control.create_label("center", "行数", 3, 0, 10)
        self.sub1_control.create_button("書き込み", lambda:self.on_button_click(), self.bg_white, self.font_black, 100, 100, 50, 30)
        # self.sub1_control.create_button("Excel開く", lambda:self.open_excel(), self.bg_white, self.font_black, 150, 100, 50, 30)
        
    def create_sub2(self):
        print('create sub2')
        sub_win2 = Toplevel(self.root)
        self.sub2_control = WindowController(sub_win2, "value_table", self.bg_white, "250x200")
        self.sub2_control.create_label2("center", "シート", 0, 0, )
        self.sub2_control.create_combobox(len(self.ws_list), 10, "center", self.ws_list, 0, 1)
        self.alphabet_cell = self.sub2_control.create_label("center", "基準セル", 1, 0, 5)
        self.number_cell = self.sub2_control.create_entry(5, 1, 2 )
        self.number_column = self.sub2_control.create_label("center", "列数", 2, 0, 10)
        self.number_row = self.sub2_control.create_label("center", "行数", 3, 0, 10)
        self.sub2_control.create_button("書き込み", lambda:self.on_button_click(), self.bg_white, self.font_black, 100, 100, 50, 30)
        # self.sub2_control.create_button("Excel開く", lambda:self.open_excel(), self.bg_white, self.font_black, 150, 100, 50, 30)
    
    def on_button_click(self):
        print('on_button')
        self.getAlphabet_cell = self.alphabet_cell.get()
        self.getNumber_cell = int(self.number_cell.get())
        self.getNumber_column = int(self.number_column.get())
        self.getNumber_row = int(self.number_row.get())
        self.input_check_table()
        
       
    
    def create_dropdown(self):
        dv = DataValidation(
            type = "list",
            formula1 = '"〇"',
            allow_blank = True,
            showErrorMessage = True,
            errorStyle = "warning",
            errorTitle = "error",
            error = "続けますか？"
        )
        self.ws.add_data_validation(dv)

        # 現在のセル位置の文字列座標を取得
        cell_coord_start = self.ws.cell(row=self.ref_row + 2, column=self.column_number + 1).coordinate
        cell_coord_end = self.ws.cell(row=self.ref_row + self.input_rows + 1, column=self.column_number + self.input_columns).coordinate
        print(cell_coord_start,':',cell_coord_end)
        dv.add(f'{cell_coord_start}:{cell_coord_end}')
        self.ws.add_data_validation(dv)
        
        
    def create_SUM_func(self, result, target_value, start_pos, end_pos):
        print(f'レート{result}, 目標値{target_value}, 開始セル{start_pos}, 終了セル{end_pos}')
        row_result, column_result = result
        row_target_value, column_target_value = target_value
        row_start_pos, column_start_pos = start_pos
        row_end_pos, column_end_pos = end_pos
        #数値座標をアルファベット座標に変換　例(1,1) → A1
        cell_result = self.ws.cell(row=row_result, column=column_result).coordinate
        cell_target_value = self.ws.cell(row=row_target_value, column=column_target_value).coordinate
        cell_start = self.ws.cell(row=row_start_pos, column=column_start_pos).coordinate
        cell_end = self.ws.cell(row=row_end_pos, column=column_end_pos).coordinate
        print(f'レート{cell_result}, 目標値{cell_target_value}, 開始セル{cell_start}, 終了セル{cell_end}')
        #Excelに計算式を貼り付け(Excelの計算式と同じ書き方でできる)
        self.ws[f'{cell_result}'].value = f'=IF({cell_target_value}=0,"0.0%",TEXT(SUM({cell_start}:{cell_end})/{cell_target_value}*100,"0.0")&"%")' 
        
    
    def sub1_create_table(self):
        line = openpyxl.styles.Side(style="thin", color="000000")     #普通線・黒色
        border = openpyxl.styles.Border(top=line, bottom=line, left=line, right=line)      #上下左右を線に適応
        self.ws.column_dimensions[self.ref_column].width = 14
        for j in range(0, self.input_rows + 2):    #縦軸に枠線を描画  input_rows + deadDay + taskline
            self.row = self.ref_row + j
            for self.index in range(0, self.input_columns + self.cell_employeeName): #横軸に枠線を描画  
                print('sub1')
                cell = self.ws.cell(self.row, self.column_number + self.index)
                cell.border = border
                cell.alignment = Alignment(horizontal = 'center',
                                           vertical = 'center',
                                           wrap_text = False)
                    
                if self.deadDay_row == self.row:
                    # deadDay_line
                    deadDay_color = openpyxl.styles.PatternFill(fgColor="a7f542", bgColor="a7f542", fill_type="solid")
                    cell.fill = deadDay_color
                    if self.column_number == self.column_number + self.index:
                        cell.value = "deadDay"

                elif self.task_row == self.row:
                    # task_line
                    task_color = openpyxl.styles.PatternFill(fgColor="fed5f8", bgColor="fed5f8", fill_type="solid")
                    cell.fill = task_color
                    if self.column_number == self.column_number + self.index:
                        cell.value = "task"
                        
                elif self.column_number == self.column_number + self.index:
                    # employee_line
                    employee_color = openpyxl.styles.PatternFill(fgColor="ffe3c0", bgColor="ffe3c0", fill_type="solid")
                    cell.fill = employee_color           
                        
                else:
         
                    pass
                
        self.create_dropdown()   
        self.save_comp = self.sub1_control.create_label3("center", "Excel書き込み完了", 25, 150, 200, 30)
        self.wb.save(file_path)
        print("save  sub1 complete")
        
        
    def sub2_create_table(self):
        line = openpyxl.styles.Side(style="thin", color="000000")     #普通線・黒色
        border = openpyxl.styles.Border(top=line, bottom=line, left=line, right=line)      #上下左右を線に適応
        self.ws.column_dimensions[self.ref_column].width = 14
        for j in range(0, self.input_rows + 2):    #縦軸に枠線を描画  input_rows + deadline + taskline
            self.row = self.ref_row + j
            for self.index in range(0, self.input_columns + self.cell_employeeName + 2): #横軸に枠線を描画  
                
                cell = self.ws.cell(self.row, self.column_number + self.index)
                cell.border = border
                cell.alignment = Alignment(horizontal = 'center',
                                           vertical = 'center',
                                           wrap_text = False)
                    
                if self.deadDay_row == self.row:
                    # deadDay_line
                    deadDay_color = openpyxl.styles.PatternFill(fgColor="a7f542", bgColor="a7f542", fill_type="solid")
                    cell.fill = deadDay_color
                    if self.column_number == self.column_number + self.index:
                        cell.value = "deadDay"

                elif self.task_row == self.row:
                    # task_line
                    task_color = openpyxl.styles.PatternFill(fgColor="fed5f8", bgColor="fed5f8", fill_type="solid")
                    cell.fill = task_color
                    if self.column_number == self.column_number + self.index:
                        cell.value = "task"
                        
                elif self.column_number == self.column_number + self.index:
                    # employee_line
                    employee_color = openpyxl.styles.PatternFill(fgColor="ffe3c0", bgColor="ffe3c0", fill_type="solid")
                    cell.fill = employee_color           
                
                
                elif self.column_number + self.cell_employeeName == self.column_number + self.index:
                    #achievement_rate_line
                    achievement_rate_color = openpyxl.styles.PatternFill(fgColor="cef4ff", bgColor="cef4ff", fill_type="solid")
                    cell.fill = achievement_rate_color
                    self.create_SUM_func((self.row, self.column_number + self.cell_employeeName),
                                         (self.row, self.column_number + self.cell_achievement_rate),
                                         (self.row, self.column_number + self.cell_target_value),
                                         (self.row, self.column_number + self.cell_target_value + self.input_columns))                     
                
                elif self.column_number + self.cell_achievement_rate == self.column_number + self.index:
                    #target_value_line
                    achievement_rate_color = openpyxl.styles.PatternFill(fgColor="f9f1a8", bgColor="f9f1a8", fill_type="solid")
                    cell.fill = achievement_rate_color    
                
                else:
         
                    pass
        
          
        self.wb.save(file_path)
        self.save_comp = self.sub2_control.create_label3("center", "Excel書き込み完了", 25, 150, 200, 30)
        print("save sub2 complete")
        
if len(sys.argv) > 1:
    file_path = sys.argv[1]
    pyperclip.copy(file_path)
    print(f"ファイルパスがクリップボードにコピーされました: {file_path}")  
    time.sleep(1)
    root = tkinter.Tk()
    app = MainApp(root)
else:
    print("ファイルパスが指定されていません。")
    time.sleep(3)
    sys.exit()

    
root.mainloop()



