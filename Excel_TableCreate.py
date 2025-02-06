import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from pathlib import Path
from openpyxl import load_workbook
import string
from openpyxl.styles.alignment import Alignment
import openpyxl.styles
from tkinter import IntVar, Toplevel, ttk
import tkinter
import tkinter as tk
import pyperclip
import sys
import time
import tkinter.font as f



class WindowController:
    def __init__(self, master, title, bg_color, win_size):
        self.master = master
        self.master.title(title)
        self.master.config(bg=bg_color)
        self.master.geometry(win_size)
        
       
    def create_button(self, text, command, bg_color, font_color, xpos, ypos, width, height):
        button = tk.Button(self.master, text=text, command=command, bg=bg_color, fg=font_color)
        button.place(x=xpos, y=ypos, width=width, height=height)
        
        
    def create_labelframe(self, text, bg_color, font_color, xpos, ypos, width, height, label_title):
        self.labelframe = tk.LabelFrame(self.master, text=label_title, bg=bg_color, fg=font_color)
        self.labelframe.place(x=xpos-10, y=ypos-20, width=width+30, height=height+30)
        label = tk.Label(self.master, text=text, bg=bg_color, fg=font_color)
        label.place(x=xpos, y=ypos, width=width, height=height)
 
    
    def create_label(self, justify_pos, text, row, column, width):
        font1 = f.Font(family="Lucida Console", weight="bold", size=12, slant="italic")
        label = tk.Label(self.master, justify=justify_pos, text=text, bg="white")
        label["font"] = font1
        label.grid(row=row, column=column)
        entry = tk.Entry(self.master, width=width, justify=justify_pos)
        entry.grid(row=row , column=column + 1)
        return entry
    
    
    def create_entry(self, width, row, column):
        entry2 = tk.Entry(self.master, width=width)
        entry2.grid(row=row , column=column)
        return entry2
    

 
class MainApp:
    def __init__(self, root):
        # self.get_path()
        self.input_flag = False
        self.input_data() 
        self.root = root  
        self.bg_white = "white"
        self.font_black = "black"
        win_control = WindowController(root, "create_table", self.bg_white, "360x150")
        check_button = win_control.create_button("check", self.open_win1, self.bg_white, self.font_black, 40, 20, 100, 50)
        value_button = win_control.create_button("value", self.open_win2, self.bg_white, self.font_black, 210, 20, 100, 50)
        win_control.create_labelframe(self.file_path, self.bg_white, self.font_black, 20, 100, 300, 20, "File_Path")
        # self.sub_win1.create_combobox(sub_win1, combo_list)
       
        # self.get_data()
     
        if self.input_flag == True:
            self.check_table()
            self.create_table()
        
        
    def get_path(self):
        if len(sys.argv) > 1:
            self.file_path = sys.argv[1]
            pyperclip.copy(self.file_path)
            print(f"ファイルパスがクリップボードにコピーされました: {self.file_path}")
            
        else:
            print("ファイルパスが指定されていません。")
            
   
    
   
    def input_data(self):
        self.file_path = Path(r"C:\Users\yuta_\OneDrive\デスクトップ\aaa.xlsx")
        self.wb = load_workbook(self.file_path, data_only=True)
        self.ws_list = self.wb.sheetnames
        self.ws = self.wb.worksheets[1]
             
        
    def check_table(self):
        self.ref_column = self.getAlphabet_cell    #基準横軸
        self.ref_row = self.getNumber_cell     #基準縦軸
        self.deadline_row = self.ref_row
        self.task_row = self.ref_row + 1
        self.employee_column = self.ref_column
        
        self.reference_cell = [self.ref_column + str(self.ref_row)]
        # self.reference_cell = self.ref_cell
        self.input_columns = self.getNumber_column   #横軸   
        self.input_rows = self.getNumber_row       #縦軸
        self.cell_employee = 1   #社員ｺｰﾄﾞ用ｾﾙ
        self.cell_task = 2  #期限用ｾﾙ + ﾀｽｸ用ｾﾙ  
        return self.ref_column
     
     
    def open_win1(self):
        self.check_sub = self.create_sub1()

    def open_win2(self):
        self.value_sub = self.create_sub2()


    def create_sub1(self):
        sub_win1 = Toplevel(self.root)
        self.sub1_control = WindowController(sub_win1, "check_table", self.bg_white, "250x200")
        self.sub1_control.create_label("center", "シート", 0, 0, 10)
        self.alphabet_cell = self.sub1_control.create_label("center", "基準セル", 1, 0, 5)
        self.number_cell = self.sub1_control.create_entry(5, 1, 2 )
        self.number_column = self.sub1_control.create_label("center", "列数", 2, 0, 10)
        self.number_row = self.sub1_control.create_label("center", "行数", 3, 0, 10)
        self.sub1_control.create_button("書き込み", lambda:self.on_button_click(), self.bg_white, self.font_black, 100, 100, 50, 30)
   
   
    def create_sub2(self):
        sub_win2 = Toplevel(self.root)
        self.sub2_control = WindowController(sub_win2, "value_table", self.bg_white, "300x200")
    
    
    def on_button_click(self):
       self.getAlphabet_cell = self.alphabet_cell.get()
       self.getNumber_cell = int(self.number_cell.get())
       self.getNumber_column = int(self.number_column.get())
       self.getNumber_row = int(self.number_row.get())
       self.input_flag = True     #ﾎﾞﾀﾝonのﾌﾗｸﾞ
       print("button_on")
       
   
       
    def create_table(self):
        line = openpyxl.styles.Side(style="thin", color="000000")     #普通線・黒色
        border = openpyxl.styles.Border(top=line, bottom=line, left=line, right=line)      #上下左右を線に適応
        self.ws.column_dimensions[self.ref_column].width = 14
        
        number = ord(self.ref_column.upper()) - ord('A') + 1
        for j in range(0, self.input_rows + self.cell_task):    #縦軸に枠線を描画
            count_row = self.ref_row + j
            for i in string.ascii_uppercase[number-1 : (number-1) + self.input_columns + self.cell_employee]:     #横軸に枠線を描画
                print(i + str(count_row))
                cell = self.ws[i + str(count_row)]
                cell.border = border
                cell.alignment = Alignment(horizontal = 'center',
                                        vertical = 'center',
                                        wrap_text = False)
                
                if self.employee_column == i:
                    print('employee')   
                    employee_color = openpyxl.styles.PatternFill(fgColor="B8CCE4", bgColor="FCE4D6", fill_type="darkGrid")
                    cell.fill = employee_color
                    
                else:
                    pass
                
                if self.deadline_row == count_row:
                    print('color1')
                    deadline_color = openpyxl.styles.PatternFill(fgColor="99FF00", bgColor="99FF00", fill_type="solid")
                    cell.fill = deadline_color
                else:
                    pass
            
                if self.task_row == count_row:
                    print('color2')
                    task_color = openpyxl.styles.PatternFill(fgColor="B8CCE4", bgColor="FFE4D6", fill_type="gray0625")
                    cell.fill = task_color
                else:
                    pass
    
        self.wb.save(self.file_path)
        print('save')
    
        
if __name__ == "__main__":


    root = tkinter.Tk()
    app = MainApp(root)
    
    root.mainloop()